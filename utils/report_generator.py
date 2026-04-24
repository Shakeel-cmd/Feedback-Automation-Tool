import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

from config.settings import GREEN_HEX, FOOTER_TEXT, EXCEL_HEADERS, COL_WIDTHS, ZOOM_SCALE


def clean_text(s) -> str:
    if pd.isna(s):
        return ""
    s = str(s).replace("\xa0", " ").replace("\t", " ")
    return re.sub(r"\s+", " ", s).strip()


def safe_filename(s) -> str:
    return re.sub(r'[\/\\\:\*\?\"\<\>\|]', "_", clean_text(s)) or "Unknown"


def fmt_date(v) -> str:
    if pd.isna(v) or v == "":
        return ""
    try:
        return pd.to_datetime(v).strftime("%d %B %Y")
    except Exception:
        return str(v)


def unique_path(folder: Path, base: str, ext: str = ".xlsx") -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    i = 1
    target = folder / (base + ext)
    while target.exists():
        target = folder / f"{base} ({i}){ext}"
        i += 1
    return target


def get_month_year_folder(date_str: str) -> str:
    try:
        d = pd.to_datetime(date_str)
        return d.strftime("%B %Y")
    except Exception:
        return "Unknown Period"


def generate_report(grp: pd.DataFrame, course: str, pl: str, date_str: str, out_folder: Path) -> tuple:
    """
    Builds and saves one feedback Excel report for a session group.
    Returns (file_path: Path, ratings: list[float]).

    Columns: A=Sr No, B=Best Part, C=Rating, D=Improvement
    """
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill = PatternFill(start_color=GREEN_HEX, end_color=GREEN_HEX, fill_type="solid")
    font_header = Font(name="Arial", size=14, bold=True)
    font_sub = Font(name="Arial", size=12, bold=True)
    font_body = Font(name="Arial", size=12)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Report"

    # Row 1 — report title (spans A:D)
    ws.merge_cells("A1:D1")
    topic_val = grp["Topic"].dropna()
    ws["A1"].value = f"Feedback Report - {topic_val.iloc[0] if len(topic_val) > 0 else course}"
    ws["A1"].font = font_header
    ws["A1"].fill = green_fill
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 37.5

    # Row 2 — date (spans A:D)
    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Date: {date_str}"
    ws["A2"].font = font_sub
    ws["A2"].alignment = align_center
    ws.row_dimensions[2].height = 18.75

    # Row 3 — column headers (4 columns via EXCEL_HEADERS)
    for ci, h in enumerate(EXCEL_HEADERS, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = font_sub
        c.fill = green_fill
        c.alignment = align_center
        c.border = border
    ws.row_dimensions[3].height = 30.95

    # Build data tuples
    excel_rows = []
    for idx, row in enumerate(grp.itertuples(), start=1):
        try:
            rv = float(row.Rating) if row.Rating not in ("", "0", "None") else None
        except Exception:
            rv = None
        imp = str(row.Improvement).strip()
        excel_rows.append((idx, row.BestPart, rv, imp))

    # Data rows starting at row 4
    # Col 1=Sr No, 2=Best Part, 3=Rating, 4=Improvement
    r = 4
    ratings = []
    for sr_no, bp, rv, imp in excel_rows:
        ws.cell(r, 1, sr_no)
        ws.cell(r, 2, bp)
        ws.cell(r, 3, rv)
        if rv:
            ratings.append(rv)

        # Improvement column (D = col 4)
        cD = ws.cell(r, 4)
        if imp in ("", "0", "-", "None"):
            cD.value = "No comments from the Learner"
            cD.font = Font(name="Arial", size=12, bold=True, color="FF0000")
        else:
            cD.value = imp
            cD.font = font_body

        for ci in range(1, 5):
            c = ws.cell(r, ci)
            c.alignment = align_center
            c.border = border
        ws.row_dimensions[r].height = 30.00
        r += 1

    # Average rating row
    avg_row = r
    ws.cell(avg_row, 2, "Average Rating").font = Font(name="Arial", size=12, bold=True)
    ws.cell(avg_row, 2).fill = green_fill
    ws.cell(avg_row, 2).alignment = align_center
    ws.cell(avg_row, 3, f'=IFERROR(AVERAGEIFS(C4:C{r-1},C4:C{r-1},">0"),"")').number_format = "0.00"
    ws.cell(avg_row, 3).font = Font(name="Arial", size=12, bold=True)
    ws.cell(avg_row, 3).fill = green_fill
    ws.cell(avg_row, 3).alignment = align_center
    ws.cell(avg_row, 4).fill = green_fill
    ws.cell(avg_row, 4).alignment = align_center
    ws.row_dimensions[avg_row].height = 25.00

    # Conditional formatting: ratings <= 3 shown in red bold
    ws.conditional_formatting.add(
        f"C4:C{r-1}",
        CellIsRule(operator="lessThanOrEqual", formula=["3"], font=Font(color="FF0000", bold=True)),
    )

    # Footer row (spans A:D)
    footer_row = avg_row + 2
    ws.merge_cells(f"A{footer_row}:D{footer_row}")
    fcell = ws[f"A{footer_row}"]
    fcell.value = FOOTER_TEXT
    fcell.font = Font(name="Arial", size=12, bold=True, italic=True, color="000000")
    fcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[footer_row].height = 25

    # Sheet settings
    ws.freeze_panes = "A4"
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.sheet_view.zoomScale = ZOOM_SCALE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Save
    base_name = f"Feedback Report - {safe_filename(course)} - {safe_filename(pl)}"
    file_path = unique_path(out_folder, base_name)
    wb.save(file_path)

    return file_path, ratings
