# app_v1_1_4.py — Feedback Automation Tool (Professional Live Dashboard Release)
# Developed by EMERITUS — Feedback Automation Tool

import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os, re
import time

# ------------------ Helper Functions ------------------
def clean_text(s):
    if pd.isna(s):
        return ""
    s = str(s).replace("\xa0", " ").replace("\t", " ")
    return re.sub(r"\s+", " ", s).strip()

def safe_filename(s):
    return re.sub(r'[\/\\\:\*\?\"\<\>\|]', "_", clean_text(s)) or "Unknown"

def fmt_date(v):
    if pd.isna(v) or v == "":
        return ""
    try:
        return pd.to_datetime(v).strftime("%d %B %Y")
    except Exception:
        return str(v)

def unique_path(folder: Path, base: str, ext: str = ".xlsx"):
    folder.mkdir(parents=True, exist_ok=True)
    i = 1
    target = folder / (base + ext)
    while target.exists():
        target = folder / f"{base} ({i}){ext}"
        i += 1
    return target
log_entries = []
def get_month_year_folder(date_str):
    """
    Converts '01 January 2025' → 'January 2025'
    """
    try:
        d = pd.to_datetime(date_str)
        return d.strftime("%B %Y")
    except:
        return "Unknown Period"

# ------------------ Streamlit Config ------------------
# ------------------ Streamlit Config ------------------
st.set_page_config(page_title="Feedback Automation Tool", page_icon=None, layout="wide")

# ---------- Simple Email Authentication (No impact on report logic) ----------
# Pre-approved users: email -> display name
ALLOWED_USERS = {
    "hariharan.v@emeritus.org".lower(): "Hariharan",
    "mohammed.shakeel@emeritus.org".lower(): "Shakeel",
    # add more here...
}

# Initialise session state flags
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["username"] = ""

def auth_screen():
    st.markdown(
        "<h2 style='text-align:center; margin-bottom:0;'>Feedback Automation Tool</h2>",
        unsafe_allow_html=True,
    )
    st.markdown("<p style='text-align:center;'>🔒 Please sign in with your official email ID</p>", unsafe_allow_html=True)

    # Center the form nicely
    col_left, col_center, col_right = st.columns([1, 2, 1])
    with col_center:
        with st.form("email_login_form", clear_on_submit=False):
            email = st.text_input("Official email ID", placeholder="your.name@company.com")
            submitted = st.form_submit_button("Enter Dashboard")

            if submitted:
                email_clean = email.strip().lower()
                if email_clean in ALLOWED_USERS:
                    st.session_state["authenticated"] = True
                    st.session_state["username"] = ALLOWED_USERS[email_clean]
                    st.success(f"Welcome, {st.session_state['username']} 👋")
                    st.rerun()  # reload app in authenticated state
                else:
                    st.error("Access denied. This email is not registered for this tool.")

# ---------- AUTH BLOCK (already working) ----------

if not st.session_state.get("authenticated", False):
    auth_screen()
    st.stop()

# ------------------ Welcome Label (Top-left in main area) ------------------
if st.session_state.get("username"):
    st.markdown(
        f"""
        <div style="
            width:100%;
            text-align:left;
            font-size:24px;
            font-weight:600;
            color:#00B050;
            margin-top:0px;
            margin-bottom:75px;
        ">
            👋 Welcome, <span style="color:white;">{st.session_state['username']}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
# ---------------------------------------------------------------------------

LOGO_PATH = Path(__file__).parent / "Emeritus.jpg"

# 🌗 Adaptive Styling for Main Content / Right Panel
# -----------------------
st.markdown("""
<style>
/* 🌞 Light Mode */
@media (prefers-color-scheme: light) {
    /* Main container */
    [data-testid="stAppViewContainer"] {
        background-color: #FFFFFF !important;
        color: black !important;
    }

    /* Headers and titles */
    .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4 {
        color: black !important;
    }

    /* Tables and logs */
    .stDataFrame, .stTable {
        background-color: #FFFFFF !important;
        color: black !important;
        border: 1px solid #ddd !important;
    }

    /* Progress bars */
    div[data-testid="stProgress"] div[role="progressbar"] {
        background-color: #00B050 !important;
    }

    /* Success / info / warning boxes */
    .stAlert {
        background-color: #F0F2F6 !important;
        color: black !important;
        border: 1px solid #ddd !important;
    }
}

/* 🌙 Dark Mode */
@media (prefers-color-scheme: dark) {
    /* Main container */
    [data-testid="stAppViewContainer"] {
        background-color: #0E1117 !important;
        color: white !important;
    }

    /* Headers and titles */
    .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4 {
        color: white !important;
    }

    /* Tables and logs */
    .stDataFrame, .stTable {
        background-color: #1E1E1E !important;
        color: white !important;
        border: 1px solid #333 !important;
    }

    /* Progress bars */
    div[data-testid="stProgress"] div[role="progressbar"] {
        background-color: #00B050 !important;
    }

    /* Success / info / warning boxes */
    .stAlert {
        background-color: #1E1E1E !important;
        color: white !important;
        border: 1px solid #333 !important;
    }
}
</style>
""", unsafe_allow_html=True)


# ------------------ Header ------------------
st.markdown(
    """
    <div style='width:100%; text-align:center; padding-top:10px; padding-bottom:6px;'>
        <h1 style='margin:0; font-size:34px; color:white;'>Feedback Automation Tool</h1>
    </div>
    <hr style='margin-top:8px; border-color:#555;'>
    """,
    unsafe_allow_html=True
)

# ------------------ Sidebar ------------------
with st.sidebar:

    st.image(str(LOGO_PATH), width=220)
    st.markdown("<h2 style='color:white;'>⚙️ Work-Flow</h2>", unsafe_allow_html=True)

    # ---------------- TEMPLATE DOWNLOAD ----------------
    template_path = Path(__file__).parent / "Automation Template.xlsx"

    st.markdown("### 📥 Download Automation Template")

    if template_path.exists():
        with open(template_path, "rb") as tfile:
            st.download_button(
                label="⬇️ Download Excel Template",
                data=tfile,
                file_name="Feedback_Automation Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="template_download_btn"
            )
    else:
        st.warning("Template file not found — please add 'Automation Template.xlsx' to the app folder.")


    # ---------------- FILE UPLOAD ----------------
    uploaded = st.file_uploader(
        "📁 Upload Automation Excel Sheet",
        type=["xlsx"]
    )

    # ---------------- LOB SELECTION ----------------
    lob_choice = st.selectbox(
        "🏷️ Select LOB",
        ["Select", "Tech Certs", "SEPO", "OC,DD,BC", "All"],
        index=0
    )

    # ---------------- OUTPUT PATH ----------------
    downloads_path = Path.home() / "Downloads"
    out_folder_input = st.text_input(
        "💾 Output folder path",
        str(downloads_path)
    )

    # ---------------- GENERATE BUTTON ----------------
    generate_btn = st.button("🚀 Generate Reports")
# ------------------ Summary Dashboard (Enhanced) ------------------
st.markdown("<h3 style='color:white; text-align:center;'>📊 Summary Dashboard</h3>", unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

with col1:
    total_reports = st.metric("Total Reports Generated", "0", delta="↗ Ready")
    #st.progress(0.0, text="Processing...")

with col2:
    avg_rating_display = st.metric("Average Rating", "0.00 ⭐", delta="No Data Yet")

with col3:
    total_lob_display = st.metric("LOBs Processed", "0", delta="Stable")

st.markdown("""
<div style='text-align:center; color:#00B050; font-style:italic;'>
Dashboard auto-refreshes after each report batch generation.
</div>
<hr style='margin-top:10px; border-color:#555;'>
""", unsafe_allow_html=True)


# ------------------ Progress Section ------------------
st.markdown("<h3 style='color:white;'>🧾 Live Progress Log</h3>", unsafe_allow_html=True)
log_placeholder = st.empty()
progress = st.progress(0)

# ------------------ Main Logic ------------------
if uploaded and generate_btn:
        # 🔁 Reset workspace for every new upload
    base_outroot = Path(out_folder_input).expanduser().resolve()
    if base_outroot.exists():
        for item in base_outroot.iterdir():
            if item.is_dir() and item.name.startswith("Feedback Reports for"):
                import shutil
                shutil.rmtree(item)
    else:
        base_outroot.mkdir(parents=True, exist_ok=True)
        
    with st.spinner("🧹 Resetting workspace..."):
        time.sleep(1)
        st.success("✅ Workspace cleaned and ready!")


    if lob_choice == "Select":
        st.warning("⚠️ Please select a valid LOB before generating reports.")
        st.stop()

    tmp_path = Path("temp_automation.xlsx")
    with open(tmp_path, "wb") as f:
        f.write(uploaded.getbuffer())

    raw = pd.read_excel(tmp_path, sheet_name="Automation", engine="openpyxl", dtype=object)
    raw.columns = [str(c).strip() for c in raw.columns]

    rows = []
    for _, r in raw.iterrows():
        rows.append({
            "Sr": clean_text(r.get("Sr No", "")),
            "Date": fmt_date(r.get("Date", "")),
            "BestPart": clean_text(r.get("What was the best part of this session and how has it helped you?", "")),
            "Rating": clean_text(r.get("Rate the overall satisfaction level of this session", "")),
            "Improvement": clean_text(r.get("What according to you could be improved in this session?", "")),
            "PL": clean_text(r.get("PL Name", "")),
            "Course": clean_text(r.get("Course Name", "")),
            "Topic": clean_text(r.get("Topic Name", "")),
            "LOB": clean_text(r.get("LOB", "")).upper(),
        })
    df = pd.DataFrame(rows)

    if lob_choice == "All":
        lob_list = sorted([l for l in df["LOB"].unique() if l])
    else:
        lob_list = [lob_choice.upper()]

    base_outroot = Path(out_folder_input).expanduser().resolve()
    base_outroot.mkdir(parents=True, exist_ok=True)

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    font_header = Font(name="Arial", size=14, bold=True)
    font_sub = Font(name="Arial", size=12, bold=True)
    font_body = Font(name="Arial", size=12)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    log_entries = []
    all_ratings = []

    groups = []
    for lob in lob_list:
        subset = df[df["LOB"] == lob]
        if not subset.empty:
            grouped = subset.groupby(["Course", "PL", "Date", "LOB"], dropna=False)
            for key, grp in grouped:
                groups.append((key, grp))

    total_groups = len(groups)
    if total_groups == 0:
        st.warning("No records found for the selected LOB.")
        st.stop()

    for i, ((course, pl, date_str, lob), grp) in enumerate(groups, start=1):
        progress.progress(int(i / total_groups * 100))

        session_folder = f"Feedback Reports for {date_str}" if date_str else "Feedback Reports"
        out_folder = base_outroot / session_folder / lob
        out_folder.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback Report"

        # Header
        ws.merge_cells("A1:D1")
        ws["A1"].value = f"Feedback Report - {grp['Topic'].dropna().iloc[0] if len(grp['Topic'].dropna())>0 else course}"
        ws["A1"].font = font_header
        ws["A1"].fill = green_fill
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 37.5

        ws.merge_cells("A2:D2")
        ws["A2"].value = f"Date: {date_str}"
        ws["A2"].font = font_sub
        ws["A2"].alignment = align_center
        ws.row_dimensions[2].height = 18.75

        headers = [
            "Sr No",
            "What was the best part of this session and how has it helped you?",
            "Rate the overall satisfaction level of this session",
            "What according to you could be improved in this session?"
        ]
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = font_sub
            c.fill = green_fill
            c.alignment = align_center
            c.border = border
        ws.row_dimensions[3].height = 30.95

        r = 4
        for idx, row in enumerate(grp.itertuples(), start=1):
            ws.cell(r, 1, idx)
            ws.cell(r, 2, row.BestPart)
            try:
                rating = float(row.Rating) if row.Rating not in ("", "0", "None") else None
            except:
                rating = None
            ws.cell(r, 3, rating)
            if rating:
                all_ratings.append(rating)

            imp = str(row.Improvement).strip()
            cD = ws.cell(r, 4)
            if imp in ("", "0", "-", "None"):
                cD.value = "No comments from the Learner"
                cD.font = Font(name="Arial", size=12, bold=True, color="FF0000")
            else:
                cD.value = imp
                cD.font = font_body
            for ci in range(1, 5):
                c = ws.cell(r, ci)
                c.alignment = align_center
                c.border = border
            ws.row_dimensions[r].height = 30.00
            r += 1

        avg_row = r
        ws.cell(avg_row, 2, "Average Rating").font = Font(name="Arial", size=12, bold=True)
        ws.cell(avg_row, 2).fill = green_fill
        ws.cell(avg_row, 2).alignment = align_center
        ws.cell(avg_row, 3, f'=IFERROR(AVERAGEIFS(C4:C{r-1},C4:C{r-1},">0"),"")').number_format = "0.00"
        ws.cell(avg_row, 3).font = Font(name="Arial", size=12, bold=True)
        ws.cell(avg_row, 3).fill = green_fill
        ws.cell(avg_row, 3).alignment = align_center
        ws.cell(avg_row, 4).fill = green_fill
        ws.cell(avg_row, 4).alignment = align_center
        ws.row_dimensions[avg_row].height = 25.00

        # Footer
        footer_row = avg_row + 2
        ws.merge_cells(f"A{footer_row}:D{footer_row}")
        fcell = ws[f"A{footer_row}"]
        fcell.value = "Developed by EMERITUS — Feedback Automation Tool"
        fcell.font = Font(name="Arial", size=12, bold=True, italic=True, color="000000")
        fcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[footer_row].height = 25
        
        # ✅ Freeze header rows and columns above A4
        ws.freeze_panes = "A4"

        # Column widths
        ws.column_dimensions["A"].width = 18.86
        ws.column_dimensions["B"].width = 42.86
        ws.column_dimensions["C"].width = 42.57
        ws.column_dimensions["D"].width = 47.86

        ws.sheet_view.zoomScale = 90
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        base_name = f"Feedback Report - {safe_filename(course)} - {safe_filename(pl)}"
        file_path = unique_path(out_folder, base_name)
        wb.save(file_path)
        # ---------------------------------------
        # LOG ENTRY (must be inside the loop)
        # ---------------------------------------
        log_entries.append({
            "S.No": i,
            "Course": course,
            "PL": pl,
            "LOB": lob,
            "Session Date": date_str,
            "File": str(file_path)
        })

        # Update Live Dashboard (inside loop)
        avg_rating_display.metric(
            "Average Rating",
            f"{(sum(all_ratings)/len(all_ratings)):.2f}" if all_ratings else "0.00"
        )
        total_reports.metric("Total Reports Generated", str(len(log_entries)))
        total_lob_display.metric("LOBs Processed", str(len(lob_list)))


    # Show Log
    df_log = pd.DataFrame(log_entries)
    df_log.index = range(1, len(df_log)+1)
    df_log.index.name = "S.No"
    log_placeholder.dataframe(df_log, height=300)

    # Download Log
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_fname = f"Feedback_Report_Log_{ts}.xlsx"
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_log.to_excel(writer, index=True)
    bio.seek(0)
    st.download_button("⬇️ Download Generation Log (Excel)", bio, log_fname,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Open Folder
    st.success(f"✅ Feedback Reports Generated Successfully for {lob_choice}")
    # ------------------------

# ------------------------------------------
# 📦 ZIP DOWNLOAD SECTION (Refined Logic)
# ------------------------------------------
# 📦 ZIP DOWNLOAD SECTION (Refined Logic - Unique Keys per Widget)
# ---------------------------------------------------------------
with st.expander("📦 Download Feedback Reports"):
    import shutil, time

    if log_entries:  # ensure reports were generated
        try:
            # Identify all session folders created during generation
            session_folders = [
                f for f in base_outroot.iterdir()
                if f.is_dir() and f.name.startswith("Feedback Reports for")
            ]

            if not session_folders:
                st.warning("No generated session folders found.")
            else:
                for session_folder in session_folders:
                    lob_folders = [f for f in session_folder.iterdir() if f.is_dir()]

                    # 🔹 Case 1 — All LOBs Selected
                    if lob_choice.upper() == "ALL":
                        zip_name = f"Feedback_Reports_All_LOBs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                        zip_path = base_outroot / zip_name
                        shutil.make_archive(str(zip_path).replace(".zip", ""), "zip", session_folder)

                        with open(zip_path, "rb") as zip_file:
                            zip_bytes = zip_file.read()

                        st.download_button(
                            label="⬇️ Download All LOB Reports (ZIP)",
                            data=zip_bytes,
                            file_name=zip_name,
                            mime="application/zip",
                            key=f"all_lobs_zip_{datetime.now().strftime('%H%M%S%f')}"
                        )
                        st.caption(f"🕒 Generated at {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
                        st.success("✅ Master ZIP for All LOBs ready for download")

                        time.sleep(2)
                        zip_path.unlink(missing_ok=True)

                    # 🔹 Case 2 — Specific LOB Selected
                    else:
                        selected_lob = lob_choice.upper()
                        found = False
                        for lob_folder in lob_folders:
                            if lob_folder.name.upper() == selected_lob:
                                found = True
                                zip_name = f"Feedback_Reports_{lob_folder.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                                zip_path = base_outroot / zip_name

                                shutil.make_archive(str(zip_path).replace(".zip", ""), "zip", lob_folder)
                                with open(zip_path, "rb") as zip_file:
                                    zip_bytes = zip_file.read()

                                st.download_button(
                                    label=f"⬇️ Download {lob_folder.name} Reports (ZIP)",
                                    data=zip_bytes,
                                    file_name=zip_name,
                                    mime="application/zip",
                                    key=f"{lob_folder.name.lower()}_{datetime.now().strftime('%H%M%S%f')}_zip"
                                )
                                st.caption(f"🕒 Generated at {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
                                st.success(f"✅ ZIP ready for {lob_folder.name}")

                                time.sleep(2)
                                zip_path.unlink(missing_ok=True)
                                break

                        if not found:
                            st.warning(f"No reports found for selected LOB: {lob_choice}")
        except Exception as e:
            st.error(f"Error while creating ZIP files: {e}")
    else:
        st.info("ℹ️ Generate reports first to enable ZIP downloads.")

# ------------------------------------------
# Footer
# ------------------ Footer ------------------
st.markdown("""
<hr>
<div style='text-align:center; background-color:#2B2B2B; color:white; padding:12px; border-radius:5px;'>
<b>Developed by EMERITUS - Tech Certs (INDIA APAC)</b><br>
Version 1.1.4 | © 2025 All Rights Reserved
</div>
""", unsafe_allow_html=True)

